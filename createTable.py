import numpy as np
from openpyxl import Workbook

# Генерация случайных объемов поставок и потребностей
supplies = np.random.randint(50, 200, size=(100,))
demands = np.random.randint(1, 20, size=(1000,))

# Генерация случайных стоимостей перевозок
costs = np.random.randint(10, 100, size=(100, 1000))

# Создание файла Excel и запись данных
wb = Workbook()
ws = wb.active

# Запись объемов поставок
for i, supply in enumerate(supplies, start=1):
    ws.cell(row=i, column=1, value=supply)

# Запись объемов потребностей
for j, demand in enumerate(demands, start=1):
    ws.cell(row=1, column=j + 1, value=demand)

# Запись стоимостей перевозок
for i in range(100):
    for j in range(1000):
        ws.cell(row=i + 2, column=j + 2, value=costs[i, j])

# Сохранение данных в файл Excel
wb.save("options/transportation_data.xlsx")
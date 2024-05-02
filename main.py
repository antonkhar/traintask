import pandas as pd
import numpy as np
from scipy.optimize import linear_sum_assignment
from openpyxl import load_workbook


# Загрузка данных из файла Excel
wb = load_workbook("options/transportation_data.xlsx")
ws = wb.active

# Чтение данных из файла Excel
supplies = [ws.cell(row=i, column=1).value for i in range(1, 101)]
demands = [ws.cell(row=1, column=j + 1).value for j in range(1000)]
costs = np.array([[ws.cell(row=i + 2, column=j + 2).value for j in range(1000)] for i in range(100)])

# Решение задачи о назначениях (как частный случай транспортной задачи)
row_ind, col_ind = linear_sum_assignment(costs)

# Создание выходной матрицы
output_matrix = np.zeros_like(costs)
output_matrix[row_ind, col_ind] = 1

# Стоимость оптимального решения
optimal_cost = costs[row_ind, col_ind].sum()

# Стоимость самого дорогого решения
max_cost = np.max(costs)

# Сохранение результата в Excel
df = pd.DataFrame(output_matrix, index=range(1, 101), columns=range(1, 1001))
df.to_excel("options/transportation_result.xlsx")

# Создание и запись данных в текстовый файл
with open("options/transportation_paths.txt", "w") as file:
    for i in range(100):
        for j in range(1000):
            if output_matrix[i, j] == 1:
                file.write(f"Со склада {i+1} везем к покупателю {j+1}\n")

    file.write("\nOptimal Cost: " + str(optimal_cost) + "\n")
    file.write("Max Cost: " + str(max_cost) + "\n")

print("Optimal Cost:", optimal_cost)
print("Max Cost:", max_cost)
